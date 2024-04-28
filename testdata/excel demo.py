import openpyxl

book = openpyxl.load_workbook("C:\\Users\\sumit\\OneDrive\\Documents\\Pythondemo.xlsx")
sheet = book.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
#
# new_entry = sheet.cell(row=2, column=2).value = "Sumit"
# print(new_entry)
#
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['A2'].value)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)
